# app/file_parser.py
"""
File parsing utilities for extracting text from various document formats.

Supports:
- Text files (.txt)
- Microsoft Word (.docx)
- Excel spreadsheets (.xlsx, .xls)
- PDF documents (.pdf)
"""

import os
from typing import Optional, Tuple
import logging

def parse_file(filepath: str) -> Tuple[Optional[str], Optional[str]]:
    """
    Parse a file and extract its text content.
    
    Args:
        filepath: Path to the file to parse
        
    Returns:
        Tuple of (text_content, error_message)
        - text_content: Extracted text, or None if parsing failed
        - error_message: Error message if parsing failed, or None if successful
    """
    if not os.path.exists(filepath):
        return None, f"File not found: {filepath}"
    
    file_ext = os.path.splitext(filepath)[1].lower()
    
    try:
        if file_ext == '.txt':
            return _parse_txt(filepath), None
        elif file_ext in ['.docx']:
            return _parse_docx(filepath), None
        elif file_ext in ['.xlsx', '.xls']:
            return _parse_excel(filepath), None
        elif file_ext == '.pdf':
            return _parse_pdf(filepath), None
        else:
            return None, f"Unsupported file type: {file_ext}"
    except Exception as e:
        logging.error(f"Error parsing file {filepath}: {e}", exc_info=True)
        return None, f"Error parsing file: {str(e)}"

def _parse_txt(filepath: str) -> str:
    """Parse plain text file."""
    with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
        return f.read()

def _parse_docx(filepath: str) -> str:
    """Parse Microsoft Word .docx file."""
    from docx import Document
    doc = Document(filepath)
    paragraphs = []
    for para in doc.paragraphs:
        if para.text.strip():
            paragraphs.append(para.text)
    return '\n'.join(paragraphs)

def _parse_excel(filepath: str) -> str:
    """Parse Excel .xlsx or .xls file."""
    from openpyxl import load_workbook
    wb = load_workbook(filepath, data_only=True)
    content = []
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        content.append(f"Sheet: {sheet_name}")
        
        for row in sheet.iter_rows(values_only=True):
            row_text = []
            for cell in row:
                if cell is not None:
                    row_text.append(str(cell))
            if row_text:
                content.append('\t'.join(row_text))
        content.append('')  # Empty line between sheets
    
    return '\n'.join(content)

def _parse_pdf(filepath: str) -> str:
    """Parse PDF file."""
    import PyPDF2
    text = []
    with open(filepath, 'rb') as f:
        pdf_reader = PyPDF2.PdfReader(f)
        for page_num, page in enumerate(pdf_reader.pages, 1):
            try:
                page_text = page.extract_text()
                if page_text.strip():
                    text.append(f"Page {page_num}:\n{page_text}")
            except Exception as e:
                logging.warning(f"Error extracting text from PDF page {page_num}: {e}")
    
    return '\n\n'.join(text)


